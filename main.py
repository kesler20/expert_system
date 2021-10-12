from flask import redirect, url_for, render_template, request, session, send_from_directory, flash
import os
import threading
import sqlalchemy
import datetime
import openpyxl as xl
from database_models import *

#------------------------------------FLASK APPLICATION-------------------------------------------------
@app.route('/static/<path:filename>')
def serve_static(filename):
    global ROOT_DIR

    return send_from_directory(os.path.join(ROOT_DIR, 'static', 'js'),   filename)    
@app.route('/logout', methods=['POST', 'GET'])
def logout():
    if 'username' in session:
        flash('You have been logged out!!')
        reset()
    else:
        pass
    return render_template('index.html')

@app.route('/login', methods=['POST', 'GET'])
def login():
    if request.method == 'POST':

        db.create_all()
        session.permanent = True

        date = request.form['date']
        node = request.form['Node']
        participant_username = request.form['participants']

        #store it in the database if new
        try:
            found_user = UserAccount.query.filter_by(username=participant_username).first()    
        except sqlalchemy.exc.OperationalError:
            found_user = False
        if found_user:
            session['username'] = found_user.username 
            session['Node'] = request.form['Node']
            session['date'] = date
            session['participants'] = participant_username
        else:
            user = UserAccount(username=participant_username, participant_id=increment_id())
            print(user)
            post = Post(content=node, user_id=user.participant_id)#since the id is our primary key it will be assigned automatically
            print(post)
            db.session.add(user)
            db.session.add(post)
            db.session.commit()
            
        #store it on the session 
        if 'username' in session:
            pass
        else:
            session['date'] = request.form['date']
            session['username'] = request.form['participants']
            session['Node'] = request.form['Node']

        return render_template('login.html',todays_date=date, participants=participant_username)
    # in the event the request method is a get request
    else:
        if 'username' in session:
            date = session['date']
            participant_username = session['username']
            flash(f'Hello {participant_username}!!')
            return render_template('login.html', todays_date=date, participants=participant_username)
        else:
            flash(f'Hello please enter your details!!')
            return render_template('login.html', todays_date=datetime.datetime.now())
    

@app.route("/")
def render_home_page():
    return render_template('index.html')

@app.route('/simulation', methods=['POST', 'GET'])
def simulation():
    if request.method == 'POST':
        return redirect(url_for('data_analysis'))
    else:
        flash('Please enter data')
        return render_template('simulation.html')

@app.route("/<page>/")
def render_page(page):
    if page == 'home':
        return render_template('index.html')
    
    elif page == 'favicon.ico':
        pass

    elif page == 'index.html':
        print('start------------------------------')
        return render_template('index.html')

    elif page == 'ExpertSystem':
        return redirect(url_for('expert'))

    elif page == 'fault_tree':
        #import GUI_process_safety
        return render_template('simulation.html')

    elif page == 'web_crawler':
        import web_crawler
        return render_template('chemicals_edit.html')

    elif page == 'data_upload':
        upload_file()
        return render_template('simulation.html')

    else:
        return render_template(r'{}.html'.format(page))

@app.route('/console_database', methods=['POST', 'GET'])
def console_database():
    if request.method == 'POST':
        lel_kg = int(request.form['lel-kg'])
        lel_v = int(request.form['lel-v'])
        import test_ExpertSystem as ts
        lel = ts.lower_explosible_limit_calculations(lel_v,lel_kg)
        flash(f'Lower explosible limit: {lel}')

        #rho = request.form['rho']
        #epsilon0 = request.form['epsilon0']
        #epsionR = request.form['epsionR']
        #response = charge_dissipation(epsilon0, epsionR, rho)
        #flash(response)

        return render_template('console_database.html')
    else:
        flash('result: ')
        return render_template('console_database.html')

@app.route('/data_analysis', methods=['POST','GET'])
def data_analysis():
    if request.method == 'POST':
        filename = request.form['simulation-filename']
        X = request.form['simulation-X']
        Y = request.form['simulation-Y']
        Z = request.form['simulation-Z']
        
        all_data = db.session.query(SessionDataModel).all()
        sessdt = update_session_data(len(all_data)-1,filename, X, Y, Z)
        print(sessdt) 
        import simulation  
        return render_template('data_analysis.html')
    else:
        return render_template('data_analysis.html')

@app.route('/data_analysis2', methods=['POST','GET'])
def data_analysis2():
    
    if request.method == 'POST':
        import Event_tree_backend
        return render_template('data_analysis2.html')
    else:
        return render_template('data_analysis2.html')

@app.route("/ExpertSystem")
def expert():
    try:
        if 'username' in session:
            date = session['date']
            node = session['Node']
            participant_username = session['username']
        else:
            date = ''
            all_posts = db.session.query(Post).all()
            last_inspected_node = all_posts[len(all_posts)-1].content
            node = last_inspected_node
            participant_username = 'Please Enter'

        wb = xl.load_workbook('Expert_System.xlsm')
        ws = wb['User Information']
        date_cell = ws['C8']
        node_cell = ws['C9']
        participant_cell = ws['C10']

        if date == '':
            date_cell.value = datetime.datetime.now()
        else:
            date_cell.value = date
        node_cell.value = node
        participant_cell.value = participant_username
        wb.save('Expert_System_Session.xlsx')
        os.system('start Expert_System_Session.xlsx')
    except PermissionError:
        return '''
        <h3>Please close the Expert_System_Session.xlsx and Expert_System.xlsm files</h3>
        <h4> Then Logout and try the Login again</h4>
        '''
    except UserWarning:
        t = threading.Thread(target=initialize_web_app)
        t.start()
    
    os.system('start Expert_System_Session.xlsx')
    return redirect(url_for('render_home_page'))


#----------------------------------- COMMAND TO START-----------------------------------------------
if __name__ == '__main__':
    app.run(debug=True, port=5500)
     
    
    

